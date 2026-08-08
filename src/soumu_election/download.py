#!/usr/bin/env python3
# v1.2.4: 参院市区町村ハブ sangiinN_3_14.html / サブページ _3_14_XX 対応（第21回）
# v1.2.3: 参院 index_1.html（本結果）を優先選択（index.html が別速報の場合がある）
# v1.2.2: 参院市区町村ハブを _7/_8 両対応、indexの「市区町村別得票」リンクも利用、404をスキップ
# v1.2.1: 合同選挙区を district 分類（衆院「○○選挙区」比例と区別）
# v1.2.0: 参院市区町村を sangiinN_8.html ツリーから取得（shikuchouson フォールバック）
# v1.1.0: --chamber sangiin 対応（参院 index 取得・項番マップ）
"""Download official MIC election workbooks and convert them to JSON.

Supports House (shugiin) and House of Councillors (sangiin) index pages.
The source workbooks are always retained. Normalized output deliberately keeps
the labels printed by MIC, so name correction can be performed later.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


BASE_URL = "https://www.soumu.go.jp"
USER_AGENT = "Project-NewsDB election-results importer/1.0"

PR_BLOCK_BY_PREFECTURE = {
    "北海道": "北海道選挙区",
    **{name: "東北選挙区" for name in ("青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県")},
    **{name: "北関東選挙区" for name in ("茨城県", "栃木県", "群馬県", "埼玉県")},
    **{name: "南関東選挙区" for name in ("千葉県", "神奈川県", "山梨県")},
    "東京都": "東京都選挙区",
    **{name: "北陸信越選挙区" for name in ("新潟県", "富山県", "石川県", "福井県", "長野県")},
    **{name: "東海選挙区" for name in ("岐阜県", "静岡県", "愛知県", "三重県")},
    **{name: "近畿選挙区" for name in ("滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県")},
    **{name: "中国選挙区" for name in ("鳥取県", "島根県", "岡山県", "広島県", "山口県")},
    **{name: "四国選挙区" for name in ("徳島県", "香川県", "愛媛県", "高知県")},
    **{name: "九州選挙区" for name in ("福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県")},
}

INDEX_CODES = {
    "表紙": "00-00",
    "届出政党等別男女別新前元別候補者数（小選挙区、比例代表）": "01-01",
    "都道府県別届出政党等別新前元別候補者数（小選挙区）": "01-02",
    "小選挙区都道府県別届出政党等別新前元別候補者数": "01-02",
    "都道府県別年齢段階別候補者数（小選挙区）": "01-03",
    "小選挙区都道府県別年齢段階別候補者数": "01-03",
    "都道府県別有権者数、投票者数（小選挙区）": "02-01",
    "都道府県別有権者数、投票者数、投票率（小選挙区）": "02-01",
    "都道府県別投票率（小選挙区）": "02-01-02",
    "都道府県別有権者数、投票者数（比例代表）": "02-02",
    "都道府県別有権者数、投票者数、投票率（比例代表）": "02-02",
    "都道府県別有権者数、投票者数、投票率（うち在外）": "02-03",
    "都道府県別投票率（比例代表）": "02-02-02",
    "届出政党等別男女別新前元別当選人数（小選挙区、比例代表）": "03-01",
    "都道府県別届出政党等別新前元別当選人数（小選挙区）": "03-02",
    "小選挙区都道府県別届出政党等別新前元別当選人数": "03-02",
    "都道府県別年齢段階別当選人数（小選挙区）": "03-03",
    "小選挙区都道府県別年齢段階別当選人数": "03-03",
    "届出政党等別得票数（小選挙区）": "03-04",
    "小選挙区届出政党等別得票数": "03-04",
    "党派別得票数（比例代表）": "03-05",
    "党派別得票数（比例）": "03-05",
    "比例代表党派別得票数": "03-05",
    "都道府県別届出政党等別得票数（小選挙区）": "03-06",
    "小選挙区都道府県別届出政党等別得票数": "03-06",
    "選挙区別都道府県別党派別得票数（比例代表）": "03-07",
    "比例代表選挙区別都道府県別党派別得票数": "03-07",
    "都道府県別投票総数、有効投票数、無効投票数（小選挙区）": "03-08",
    "都道府県別選挙区別投票総数、有効投票数、無効投票数（比例代表）": "03-09",
    "選挙区別党派別得票数（比例代表）": "03-10",
    "比例代表選挙区別党派別得票数": "03-10",
    "党派別当選人数（比例代表）": "03-11",
    "比例代表党派別当選人数": "03-11",
    "党派別議席配分表（比例代表）": "03-12",
    "比例代表党派別議席配分表": "03-12",
    "候補者別得票数（小選挙区）": "03-13",
    "18歳、19歳投票状況": "03-15",
    "年齢別投票状況": "03-16",
    "管理執行上問題となった事項": "04-01",
    "都道府県別有権者数、投票者数": "05-01",
    "都道府県別有権者数、投票者数、投票率": "05-01",
    "都道府県別投票率": "05-01-02",
    "罷免を可とする投票数、可としない投票数": "05-02",
    "都道府県別投票総数、有効投票数、無効投票数": "05-03",
}

# 参院は「選挙区／比例代表」表記。項番は参院indexの見出し順に独自採番。
SANGIIN_INDEX_CODES = {
    "党派別男女別新前元別候補者数（比例代表、選挙区）": "01-01",
    "都道府県別党派別新前元別候補者数（選挙区）": "01-02",
    "都道府県別有権者数、投票者数、投票率（比例代表）": "02-01",
    "都道府県別有権者数、投票者数、投票率（比例代表）（比較）": "02-01-cmp",
    "都道府県別有権者数、投票者数、投票率（選挙区）": "02-02",
    "都道府県別有権者数、投票者数、投票率（選挙区）（比較）": "02-02-cmp",
    "党派別男女別新前元別当選人数（比例代表、選挙区）": "03-01",
    "都道府県別党派別新前元別当選人数（選挙区）": "03-02",
    "党派別得票数（比例代表）": "03-03",
    "党派別得票数（選挙区）": "03-04",
    "都道府県別党派別得票数（比例代表）": "03-05",
    "都道府県別党派別得票数（選挙区）": "03-06",
    "都道府県別投票総数、有効投票数、無効投票数（比例代表）": "03-07",
    "都道府県別投票総数、有効投票数、無効投票数（選挙区）": "03-08",
    "得票順党派別得票数（比例代表）": "03-09",
    "党派別名簿登載者別得票数、当選人数（比例代表）": "03-10",
    "都道府県別党派別名簿登載者別得票数（比例代表）": "03-11",
    "党派別議席配分表（比例代表）": "03-12",
    "候補者別得票数（選挙区）": "03-13",
    "候補者別市区町村別得票数": "03-14",
    "18歳、19歳投票状況": "03-15",
    "18歳、19歳投票状況（速報）": "03-15",
    "年齢別投票状況": "03-16",
    "年齢別投票者数調": "03-16",
    "年齢別投票率の状況": "03-16",
    "比例代表党派別名簿登載者別選挙運動費用収支報告書要旨": "03-17",
    "管理執行上問題となった事項": "04-01",
}


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\u3000", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def parse_vote(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text or text in {"-", "－", "―"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def get(session: requests.Session, url: str, attempts: int = 4) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            response = session.get(url, timeout=(15, 120))
            response.raise_for_status()
            return response
        except requests.RequestException as exc:
            last_error = exc
            if attempt + 1 < attempts:
                time.sleep(2**attempt)
    raise RuntimeError(f"取得に失敗しました: {url}: {last_error}")


def discover_page(
    session: requests.Session,
    page_url: str,
    page_kind: str,
    *,
    allow_empty: bool = False,
    label_override: str | None = None,
    index_codes: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    codes = index_codes or INDEX_CODES
    response = get(session, page_url)
    # The current and historical MIC pages declare Shift_JIS. requests' guess
    # is not consistently correct, so honor the HTML declaration explicitly.
    response.encoding = response.apparent_encoding or "shift_jis"
    soup = BeautifulSoup(response.text, "html.parser")
    links: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if not re.search(r"\.(?:xlsx|xls|pdf)(?:\?.*)?$", href, re.IGNORECASE):
            continue
        url = urljoin(page_url, href)
        if url in seen:
            continue
        seen.add(url)
        link_label = clean_text(anchor.get_text(" ", strip=True))
        icon_only_link = not link_label
        if icon_only_link:
            previous = anchor.find_previous("a")
            previous_label = clean_text(previous.get_text(" ", strip=True)) if previous else None
            link_label = f"{previous_label}（別版）" if previous_label else Path(urlparse(url).path).stem
        link_label = re.sub(r"（別ウィンドウ）$", "", link_label).strip()
        label = label_override or link_label
        if page_kind == "municipality_votes":
            category = classify_municipality_category(link_label)
            if label_override:
                # 都道府県名 + リンク種別（選挙区/比例/政党別/候補者別）
                label = f"{label_override}_{link_label}" if link_label and link_label != label_override else label_override
        else:
            category = "pdf" if urlparse(url).path.lower().endswith(".pdf") else "summary"
        source_code = codes.get(label)
        if source_code is None:
            # 党派名だけのリンク（参院 03-11 名簿登載者別など）
            if page_kind == "summary" and "名簿登載者" in (page_url + label):
                source_code = codes.get("都道府県別党派別名簿登載者別得票数（比例代表）")
            for key, code in codes.items():
                if label.startswith(key) or key in label:
                    source_code = code
                    break
        links.append(
            {
                "category": category,
                "dataset": f"{label} {link_label}" if label_override else label,
                "label": label,
                "url": url,
                "source_page": page_url,
                "page_kind": page_kind,
                "source_code": "00-00-old" if icon_only_link and link_label.startswith("表紙") else source_code,
            }
        )
    if not links and not allow_empty:
        raise RuntimeError(f"Excelリンクが見つかりません: {page_url}")
    return links


def classify_municipality_category(link_label: str) -> str:
    """市区町村 Excel の種別（衆院 smd/pr、参院 district/pr-party/pr-cand）。"""
    text = (link_label or "").strip()
    if "比例" in text:
        if "政党" in text or "党派" in text:
            return "pr-party"
        if "候補" in text:
            return "pr-cand"
        return "pr"
    if "小選挙区" in text:
        return "smd"
    # 参院: 「選挙区」「合同選挙区」
    if text == "選挙区" or "合同選挙区" in text or text.startswith("選挙区"):
        return "district"
    if text in {"政党別", "党派別"} or ((("政党" in text) or ("党派" in text)) and "別" in text):
        return "pr-party"
    if "候補" in text:
        return "pr-cand"
    # 参院第21回など: 都道府県ページ上の党名リンクは名簿候補別PDF
    if any(token in text for token in ("党", "ネット", "維新", "無所属")):
        return "pr-cand"
    # 衆院: 「○○選挙区」は比例ブロック、それ以外の都道府県リンクは小選挙区
    if text.endswith("選挙区"):
        return "pr"
    return "smd"


def discover_municipality_pages(
    session: requests.Session,
    page_url: str,
    *,
    index_codes: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    direct = discover_page(session, page_url, "municipality_votes", allow_empty=True, index_codes=index_codes)
    if direct:
        return direct
    response = get(session, page_url)
    response.encoding = response.apparent_encoding or "shift_jis"
    soup = BeautifulSoup(response.text, "html.parser")
    subpages: list[tuple[str, str]] = []
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        # 衆院: shikuchouson_XX.html
        # 参院: sangiinN_7_XX / sangiinN_8_XX / sangiinN_3_14_XX.html
        if not re.search(r"(?:shikuchouson_\d+|sangiin\d+_(?:[78]|3_14)_\d+)\.html$", href):
            continue
        subpages.append((urljoin(page_url, href), clean_text(anchor.get_text(" ", strip=True)) or "unknown"))
    if not subpages:
        raise RuntimeError(f"Excelリンクも都道府県サブページも見つかりません: {page_url}")
    links: list[dict[str, str]] = []
    for index, (subpage_url, prefecture) in enumerate(subpages):
        if index:
            time.sleep(0.25)
        links.extend(discover_page(
            session, subpage_url, "municipality_votes",
            label_override=prefecture, index_codes=index_codes,
        ))
    return links


def choose_summary_index(
    session: requests.Session,
    base: str,
    *,
    index_codes: dict[str, str],
) -> tuple[str, list[dict[str, str]]]:
    """Pick the best summary index page among index.html / index_1.html."""
    candidates = [f"{base}/index.html", f"{base}/index_1.html"]
    best_url = candidates[0]
    best_links: list[dict[str, str]] = []
    best_score = -1
    for url in candidates:
        try:
            links = discover_page(session, url, "summary", index_codes=index_codes)
        except (RuntimeError, requests.RequestException):
            continue
        coded = sum(1 for item in links if item.get("source_code"))
        score = coded * 100 + len(links)
        if score > best_score:
            best_score = score
            best_url = url
            best_links = links
    if best_score < 0:
        raise RuntimeError(f"summary index を取得できません: {base}")
    return best_url, best_links


def discover(
    session: requests.Session,
    kaiji: int,
    *,
    chamber: str = "shugiin",
) -> tuple[list[str], list[dict[str, str]]]:
    if chamber not in {"shugiin", "sangiin"}:
        raise ValueError(f"unsupported chamber: {chamber}")
    base = f"{BASE_URL}/senkyo/senkyo_s/data/{chamber}{kaiji}"
    codes = SANGIIN_INDEX_CODES if chamber == "sangiin" else INDEX_CODES
    if chamber == "sangiin":
        index_url, links = choose_summary_index(session, base, index_codes=codes)
        pages = [index_url]
    else:
        pages = [f"{base}/index.html"]
        links = discover_page(session, pages[0], "summary", index_codes=codes)

    municipality_links: list[dict[str, str]] = []
    if chamber == "shugiin" and kaiji >= 45:
        municipality_page = f"{base}/shikuchouson.html"
        pages.append(municipality_page)
        municipality_links = discover_municipality_pages(session, municipality_page)
    elif chamber == "sangiin":
        # 参院: 回によって sangiinN_8.html / sangiinN_7.html / index直リンク
        candidates: list[str] = []
        try:
            index_resp = get(session, pages[0])
            index_resp.encoding = index_resp.apparent_encoding or "shift_jis"
            index_soup = BeautifulSoup(index_resp.text, "html.parser")
            for anchor in index_soup.find_all("a", href=True):
                label = clean_text(anchor.get_text(" ", strip=True))
                href = urljoin(pages[0], anchor["href"])
                if ("市区町村別得票" in label or "市区町村別" in label) and href.endswith(".html"):
                    candidates.append(href)
        except Exception:
            pass
        candidates.extend([
            f"{base}/{chamber}{kaiji}_8.html",
            f"{base}/{chamber}{kaiji}_7.html",
            f"{base}/{chamber}{kaiji}_3_14.html",
            f"{base}/shikuchouson.html",
        ])
        municipality_page = None
        seen: set[str] = set()
        for candidate in candidates:
            if candidate in seen:
                continue
            seen.add(candidate)
            try:
                probe = get(session, candidate)
            except (RuntimeError, requests.RequestException):
                continue
            if probe.status_code == 200 and len(probe.content) > 200:
                municipality_page = candidate
                break
        if municipality_page:
            pages.append(municipality_page)
            municipality_links = discover_municipality_pages(
                session, municipality_page, index_codes=codes,
            )

    counters: dict[str, int] = defaultdict(int)
    for item in municipality_links:
        counters[item["category"]] += 1
        item["source_code"] = f"03-14-{item['category']}-{counters[item['category']]:02d}"
    links.extend(municipality_links)

    for item in links:
        if not item.get("source_code") and "年齢別" in item["label"]:
            item["source_code"] = "03-16"
        if item["label"].startswith("まとめて表示する"):
            item["source_code"] = "99-00"
        if item["label"].startswith(f"{kaiji}衆結果調全体版") or "結果調全体版" in item["label"]:
            item["source_code"] = "99-00"
        item["chamber"] = chamber
    unique = {item["url"]: item for item in links}
    return pages, list(unique.values())


def safe_filename(item: dict[str, str]) -> str:
    source_name = Path(urlparse(item["url"]).path).name
    label = re.sub(r"[^0-9A-Za-zぁ-んァ-ヶ一-龠々ー]+", "_", item["label"]).strip("_")
    code = item.get("source_code") or "unclassified"
    return f"{code}_{label}_{source_name}"


def download(session: requests.Session, item: dict[str, str], raw_dir: Path, force: bool) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / safe_filename(item)
    if path.exists() and not force:
        return path
    response = get(session, item["url"])
    suffix = path.suffix.lower()
    if suffix == ".xlsx" and not response.content.startswith(b"PK"):
        raise RuntimeError(f"Excel形式ではない応答です: {item['url']}")
    if suffix == ".xls" and not response.content.startswith(bytes.fromhex("D0CF11E0")):
        raise RuntimeError(f"Excel 97-2003形式ではない応答です: {item['url']}")
    if suffix == ".pdf" and not response.content.startswith(b"%PDF"):
        raise RuntimeError(f"PDF形式ではない応答です: {item['url']}")
    temporary = path.with_suffix(path.suffix + ".part")
    temporary.write_bytes(response.content)
    temporary.replace(path)
    return path


def rows(ws: Any) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def workbook_tables(path: Path) -> list[tuple[str, list[list[Any]]]]:
    if path.suffix.lower() == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        return [
            (ws.name, [[ws.cell_value(row, column) for column in range(ws.ncols)] for row in range(ws.nrows)])
            for ws in workbook.sheets()
        ]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [(ws.title, rows(ws)) for ws in workbook.worksheets]
    finally:
        workbook.close()


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return str(value)


def workbook_to_raw_json(path: Path, source: dict[str, str], kaiji: int) -> dict[str, Any]:
    """Return every non-empty displayed cell plus sheet/merge metadata."""
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for ws in workbook.worksheets:
            cells = [
                {"cell": cell.coordinate, "row": cell.row, "column": cell.column, "value": json_value(cell.value)}
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            sheets.append(
                {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "merged_cells": [str(item) for item in ws.merged_cells.ranges],
                    "cells": cells,
                }
            )
    finally:
        workbook.close()
    return {
        "schema_version": "1.0",
        "election_kaiji": kaiji,
        "dataset": source["dataset"],
        "category": source["category"],
        "source_url": source["url"],
        "source_page": source["source_page"],
        "source_file": path.name,
        "sheets": sheets,
    }


def xls_to_raw_json(path: Path, source: dict[str, str], kaiji: int) -> dict[str, Any]:
    import xlrd

    workbook = xlrd.open_workbook(path, formatting_info=True)
    sheets = []
    for ws in workbook.sheets():
        cells = []
        for row in range(ws.nrows):
            for column in range(ws.ncols):
                cell = ws.cell(row, column)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, workbook.datemode).isoformat()
                cells.append({"cell": cell_ref_xls(row, column), "row": row + 1, "column": column + 1, "value": value})
        sheets.append({
            "name": ws.name,
            "max_row": ws.nrows,
            "max_column": ws.ncols,
            "merged_cells": [f"{cell_ref_xls(rlo, clo)}:{cell_ref_xls(rhi - 1, chi - 1)}" for rlo, rhi, clo, chi in ws.merged_cells],
            "cells": cells,
        })
    return {
        "schema_version": "1.0",
        "election_kaiji": kaiji,
        "dataset": source["dataset"],
        "category": source["category"],
        "source_url": source["url"],
        "source_page": source["source_page"],
        "source_file": path.name,
        "sheets": sheets,
    }


def cell_ref_xls(row: int, column: int) -> str:
    letters = ""
    value = column + 1
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row + 1}"


def pdf_to_raw_json(path: Path, source: dict[str, str], kaiji: int) -> dict[str, Any]:
    from pypdf import PdfReader

    reader = PdfReader(path)
    pages = [
        {"page": number, "text": page.extract_text() or ""}
        for number, page in enumerate(reader.pages, start=1)
    ]
    embedded_attachments: list[dict[str, Any]] = []
    attachments_dir = path.parent / f"{path.stem}_attachments"
    for attachment_name, payloads in reader.attachments.items():
        for index, payload in enumerate(payloads, start=1):
            safe_name = re.sub(r'[<>:"/\\|?*]+', "_", Path(attachment_name).name)
            if len(payloads) > 1:
                attachment_path = Path(safe_name)
                safe_name = f"{attachment_path.stem}_{index}{attachment_path.suffix}"
            attachments_dir.mkdir(parents=True, exist_ok=True)
            output_path = attachments_dir / safe_name
            output_path.write_bytes(payload)
            attachment_pages: list[dict[str, Any]] = []
            if output_path.suffix.lower() == ".pdf":
                attachment_reader = PdfReader(BytesIO(payload))
                attachment_pages = [
                    {"page": number, "text": page.extract_text() or ""}
                    for number, page in enumerate(attachment_reader.pages, start=1)
                ]
            embedded_attachments.append({
                "name": attachment_name,
                "file": str(output_path.relative_to(path.parent)),
                "bytes": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
                "page_count": len(attachment_pages),
                "pages": attachment_pages,
            })
    return {
        "schema_version": "1.0",
        "election_kaiji": kaiji,
        "dataset": source["dataset"],
        "category": source["category"],
        "source_code": source.get("source_code"),
        "source_url": source["url"],
        "source_page": source["source_page"],
        "source_file": path.name,
        "page_count": len(pages),
        "pages": pages,
        "embedded_attachments": embedded_attachments,
    }


def normalize_person_name(value: str) -> str:
    return re.sub(r"[\s\u3000]+", "", value)


def find_row(table: list[list[Any]], marker: str) -> int:
    for index, row in enumerate(table):
        if any(marker in (clean_text(cell) or "") for cell in row):
            return index
    raise ValueError(f"シート {marker!r} 行が見つかりません")


def find_row_any(table: list[list[Any]], markers: tuple[str, ...]) -> int:
    for index, row in enumerate(table):
        if any(marker in (clean_text(cell) or "") for cell in row for marker in markers):
            return index
    raise ValueError(f"シート {markers!r} 行が見つかりません")


def reporting_label(row: list[Any]) -> str | None:
    first = clean_text(row[0]) if row else None
    second = clean_text(row[1]) if len(row) > 1 else None
    return second or first


def reporting_label_before(row: list[Any], first_value_column: int) -> str | None:
    labels = []
    for value in row[:first_value_column]:
        text = clean_text(value)
        if text and parse_vote(value) is None:
            labels.append(text)
    return labels[-1] if labels else None


def row_type(label: str, district: str | None = None) -> str:
    if label.endswith("計") or label in {"合計", "総計"}:
        return "total"
    if district and label == district:
        return "total"
    return "reporting_unit"


def parse_smd(path: Path, source: dict[str, str], kaiji: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for sheet_name, table in workbook_tables(path):
        try:
            candidate_row = find_row(table, "候補者名")
            party_row = find_row_any(table[candidate_row:], ("党派名", "政党名", "政党等名")) + candidate_row
        except ValueError:
            continue
        candidates: list[tuple[int, str, str | None]] = []
        for column in range(1, len(table[candidate_row])):
            candidate = clean_text(table[candidate_row][column])
            if not candidate or "得票数計" in candidate or "開票率" in candidate:
                continue
            party = clean_text(table[party_row][column]) if column < len(table[party_row]) else None
            candidates.append((column, candidate, party))
        if not candidates:
            raise ValueError(f"候補者列がありません: {path.name}/{sheet_name}")
        first_candidate_column = min(item[0] for item in candidates)
        for excel_row, row in enumerate(table[party_row + 1 :], start=party_row + 2):
            label = reporting_label_before(row, first_candidate_column)
            if not label:
                continue
            for column, candidate_raw, party in candidates:
                vote = parse_vote(row[column] if column < len(row) else None)
                if vote is None:
                    continue
                result.append({
                    "election_kaiji": kaiji,
                    "contest": "smd",
                    "prefecture": source["label"],
                    "district": clean_text(sheet_name),
                    "reporting_unit": label,
                    "row_type": row_type(label, clean_text(sheet_name)),
                    "candidate": normalize_person_name(candidate_raw),
                    "candidate_raw": candidate_raw,
                    "party": party,
                    "votes": vote,
                    "source_file": path.name,
                    "source_url": source["url"],
                    "source_sheet": sheet_name,
                    "source_row": excel_row,
                })
    return result


def parse_pr(path: Path, source: dict[str, str], kaiji: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    source_area = source["label"]
    pr_block = PR_BLOCK_BY_PREFECTURE.get(source_area, source_area)
    for sheet_name, table in workbook_tables(path):
            try:
                party_marker = find_row_any(table, ("党派名", "政党名"))
                reporting_marker = find_row_any(table[party_marker:], ("開票区名", "市区町村名")) + party_marker
            except ValueError:
                # Historical workbooks may contain an auxiliary lookup sheet.
                continue
            party_names: list[tuple[int, str]] = []
            for header_index in range(party_marker, reporting_marker + 1):
                for column in range(1, len(table[header_index])):
                    party = clean_text(table[header_index][column])
                    if not party or party.isdigit() or party in {"得票数計", "開票率（％）"}:
                        continue
                    if "党派名" in party or "開票区名" in party or "開票率" in party:
                        continue
                    if column not in {item[0] for item in party_names}:
                        party_names.append((column, party))
            if not party_names:
                raise ValueError(f"政党列がありません: {path.name}/{sheet_name}")
            first_party_column = min(item[0] for item in party_names)
            for excel_row, row in enumerate(table[reporting_marker + 1 :], start=reporting_marker + 2):
                label = reporting_label_before(row, first_party_column)
                if not label:
                    continue
                for column, party in party_names:
                    vote = parse_vote(row[column] if column < len(row) else None)
                    if vote is None:
                        continue
                    result.append({
                        "election_kaiji": kaiji,
                        "contest": "pr",
                        "block": pr_block,
                        "source_area": source_area,
                        "reporting_unit": label,
                        "row_type": row_type(label),
                        "party": party,
                        "votes": vote,
                        "source_file": path.name,
                        "source_url": source["url"],
                        "source_sheet": sheet_name,
                        "source_row": excel_row,
                    })
    return result


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".part")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(path)


def validate(records: Iterable[dict[str, Any]], contest: str) -> list[str]:
    records = list(records)
    errors: list[str] = []
    if not records:
        errors.append(f"{contest}: レコードが0件です")
    for index, record in enumerate(records):
        if record["votes"] < 0:
            errors.append(f"{contest}[{index}]: 負の得票数")
        if not record.get("reporting_unit"):
            errors.append(f"{contest}[{index}]: 開票区名なし")
    return errors


def find_project_root(script: Path) -> Path:
    for directory in (script.parent, *script.parents):
        if (directory / "requirements.txt").exists():
            return directory
    return Path.cwd()


def main() -> int:
    parser = argparse.ArgumentParser(description="総務省の選挙結果を取得してJSON化（衆院/参院）")
    parser.add_argument("--kaiji", type=int, required=True, help="選挙回次（例: 衆51 / 参27）")
    parser.add_argument(
        "--chamber", choices=("shugiin", "sangiin"), default="shugiin",
        help="議院（shugiin=衆院, sangiin=参院）",
    )
    parser.add_argument("--output", type=Path, help="出力先（既定: data/<chamber><回次>）")
    parser.add_argument("--force", action="store_true", help="既存のExcelを再取得")
    args = parser.parse_args()

    project_root = find_project_root(Path(__file__).resolve())
    output = args.output or project_root / "data" / f"{args.chamber}{args.kaiji}"
    raw_dir = output / "raw"
    session = requests.Session()
    session.headers["User-Agent"] = USER_AGENT

    source_pages, sources = discover(session, args.kaiji, chamber=args.chamber)
    smd_records: list[dict[str, Any]] = []
    pr_records: list[dict[str, Any]] = []
    manifest_sources: list[dict[str, Any]] = []
    raw_json_dir = output / "raw_json"
    for number, source in enumerate(sources, start=1):
        path = download(session, source, raw_dir, args.force)
        is_pdf = path.suffix.lower() == ".pdf"
        raw_json = (
            pdf_to_raw_json(path, source, args.kaiji)
            if is_pdf
            else xls_to_raw_json(path, source, args.kaiji)
            if path.suffix.lower() == ".xls"
            else workbook_to_raw_json(path, source, args.kaiji)
        )
        raw_json["chamber"] = args.chamber
        raw_json_path = raw_json_dir / f"{path.stem}.json"
        write_json(raw_json_path, raw_json)
        records: list[dict[str, Any]] = []
        # 参院は市区町村パーサ未対応のため semantic 変換はスキップ（raw のみ）
        if args.chamber == "shugiin":
            if source["category"] == "smd":
                records = parse_smd(path, source, args.kaiji)
                smd_records.extend(records)
            elif source["category"] == "pr":
                records = parse_pr(path, source, args.kaiji)
                pr_records.extend(records)
        manifest_sources.append({
            **source,
            "file": str(path.relative_to(output)),
            "raw_json": str(raw_json_path.relative_to(output)),
            "sha256": sha256(path),
            "normalized_records": len(records),
            "raw_cells": 0 if is_pdf else sum(len(sheet["cells"]) for sheet in raw_json["sheets"]),
            "pdf_pages": raw_json.get("page_count", 0),
            "embedded_attachments": len(raw_json.get("embedded_attachments", [])),
        })
        print(
            f"[{number:02d}/{len(sources)}] {source['label']}: "
            f"{(raw_json.get('page_count', 0) if is_pdf else sum(len(sheet['cells']) for sheet in raw_json['sheets'])):,} "
            f"{'pages' if is_pdf else 'cells'}, "
            f"{len(records):,} normalized records",
            file=sys.stderr,
        )

    errors: list[str] = []
    if args.chamber == "shugiin":
        if any(source["category"] == "smd" for source in sources):
            errors.extend(validate(smd_records, "smd"))
        if any(source["category"] == "pr" for source in sources):
            errors.extend(validate(pr_records, "pr"))
    write_json(output / "smd_votes.json", smd_records)
    write_json(output / "pr_votes.json", pr_records)
    manifest = {
        "schema_version": "1.0",
        "election_type": args.chamber,
        "election_kaiji": args.kaiji,
        "source_pages": source_pages,
        "retrieved_at": datetime.now(timezone.utc).isoformat(),
        "counts": {
            "sources": len(sources),
            "summary_sources": sum(item["category"] == "summary" for item in sources),
            "smd_sources": sum(item["category"] == "smd" for item in sources),
            "pr_sources": sum(item["category"] == "pr" for item in sources),
            "pdf_sources": sum(item["category"] == "pdf" for item in sources),
            "raw_cells": sum(item["raw_cells"] for item in manifest_sources),
            "pdf_pages": sum(item["pdf_pages"] for item in manifest_sources),
            "embedded_attachments": sum(item["embedded_attachments"] for item in manifest_sources),
            "smd_records": len(smd_records),
            "pr_records": len(pr_records),
        },
        "formats": {
            extension: sum(Path(urlparse(item["url"]).path).suffix.lower() == extension for item in sources)
            for extension in (".xls", ".xlsx", ".pdf")
        },
        "normalization_policy": {
            "excel": "セル座標・結合セルを保持してraw JSON化し、対応済み表のみsemantic変換する",
            "pdf": "原本とページ別抽出テキストを保持するが、表構造を検証するまでsemantic変換しない",
            "raw_only_is_not_normalized": True,
            "sangiin_semantic": "参院は当面 raw/raw_json まで。正規化パーサは後続",
        },
        "validation": {"ok": not errors, "errors": errors},
        "unavailable_sources": (
            [] if any(item.get("source_code") == "04-01" for item in sources) else [{
                "source_code": "04-01",
                "dataset": "管理執行上問題となった事項",
                "status": "not_published",
                "reason": "総務省indexに取得可能なリンクなし",
                "source_page": source_pages[0],
            }]
        ) + ([] if any(item["category"] in {"smd", "pr"} for item in sources) else [{
            "source_code": "03-14",
            "dataset": "市区町村別得票数",
            "status": "not_published" if args.chamber == "shugiin" else "pending_or_not_on_shikuchouson",
            "reason": (
                "総務省indexに市区町村別ページまたは取得可能なファイルへのリンクなし"
                if args.chamber == "shugiin"
                else "参院の市区町村別は別ページ構成の可能性あり。index上のリンク有無を要確認"
            ),
            "source_page": source_pages[0],
        }]),
        "sources": manifest_sources,
    }
    write_json(output / "manifest.json", manifest)
    print(json.dumps(manifest["counts"], ensure_ascii=False))
    return 0 if not errors else 2


if __name__ == "__main__":
    raise SystemExit(main())

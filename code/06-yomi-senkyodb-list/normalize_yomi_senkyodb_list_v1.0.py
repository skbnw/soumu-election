# -*- coding: utf-8 -*-
"""
読売選挙DB立候補者リスト正規化
v1.0
- references/yomi-senkyoDB-list の xlsx を行単位 JSONL / parquet に固定
- 数値正本にはしない（表示・人名照合の補助ソース）
- 衆: 2005–2026（年別シート）、参: 2004–2025（sangiin + 補選は除外既定）

出力:
  output/06-yomi-senkyodb-list/YYYYMMDD_HHMM_yomi_senkyodb_candidates.jsonl
  output/06-yomi-senkyodb-list/yomi_senkyodb_candidates.jsonl（最新）
  web/data/yomi_senkyodb_candidates.parquet
  data/warehouse/parquet/yomi_senkyodb_candidates.parquet
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import duckdb
import openpyxl

REPO = Path(__file__).resolve().parents[2]
SRC_DIR = REPO / "references" / "yomi-senkyoDB-list"
SHUGIIN_XLSX = SRC_DIR / "shugiin-can-list-2005-2026.xlsx"
SANGIIN_XLSX = SRC_DIR / "sangiin-can-list-2004-2025.xlsx"
OUT_DIR = REPO / "output" / "06-yomi-senkyodb-list"
WEB_OUT = REPO / "web" / "data" / "yomi_senkyodb_candidates.parquet"
WARE_OUT = REPO / "data" / "warehouse" / "parquet" / "yomi_senkyodb_candidates.parquet"

SHUGIIN_YEAR_TO_KAIJI = {
    2005: 44, 2009: 45, 2012: 46, 2014: 47, 2017: 48, 2021: 49, 2024: 50, 2026: 51,
}
SANGIIN_YEAR_TO_KAIJI = {
    2004: 20, 2007: 21, 2010: 22, 2013: 23, 2016: 24, 2019: 25, 2022: 26, 2025: 27,
}
GOKU_MAP = {
    "鳥取・島根": "鳥取県・島根県",
    "鳥取県・島根県": "鳥取県・島根県",
    "徳島・高知": "徳島県・高知県",
    "徳島県・高知県": "徳島県・高知県",
}
YEAR_RE = re.compile(r"(20\d{2})")
DIST_NUM_RE = re.compile(r"(\d+)\s*区")


def nfkc(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def header_index(header_row: tuple) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = nfkc(cell)
        if not key or key in idx:
            continue
        idx[key] = i
    # 全角スラッシュ表記の互換
    aliases = {
        "ブ／県": "ブ/県",
        "ブ／県CD": "ブ/県CD",
    }
    for full, half in aliases.items():
        if full not in idx and half in idx:
            idx[full] = idx[half]
        if half not in idx and full in idx:
            idx[half] = idx[full]
    return idx


def cell(row: tuple, idx: dict[str, int], *names: str) -> str:
    for name in names:
        if name in idx:
            return nfkc(row[idx[name]])
    return ""


def parse_year_from_sheet(sheet_name: str) -> int | None:
    m = YEAR_RE.search(nfkc(sheet_name))
    return int(m.group(1)) if m else None


def parse_year_from_election(election_name: str) -> int | None:
    s = nfkc(election_name)
    if "一覧" in s:
        return None
    m = YEAR_RE.search(s)
    return int(m.group(1)) if m else None


def normalize_pref(raw: str) -> str | None:
    name = nfkc(raw)
    if not name or name in ("組閣",):
        return None
    return GOKU_MAP.get(name, name)


def parse_district_number(district_label: str) -> int | None:
    s = nfkc(district_label)
    m = DIST_NUM_RE.search(s)
    if m:
        return int(m.group(1))
    if s.isdigit():
        return int(s)
    return None


def parse_dob(value: object) -> str | None:
    if value is None or value == "":
        return None
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = nfkc(value)
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if not m:
        return None
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"


def classify_contest(
    chamber: str,
    district: str,
    pr_block: str,
    election_name: str,
    kind: str,
) -> str:
    """衆は小選挙区優先（重複立候補は smd）。参は県区名があれば district。"""
    dist = nfkc(district)
    pr = nfkc(pr_block)
    blob = f"{election_name} {kind}"
    if chamber == "shugiin":
        if dist and ("区" in dist or dist.isdigit() or parse_district_number(dist) is not None):
            return "smd"
        if pr or "比例" in blob:
            return "pr"
        return "unknown"
    # sangiin
    if dist and dist not in ("", "全国", "比例") and "ブ" not in dist:
        return "district"
    if pr or "比例" in blob or dist in ("全国", "比例"):
        return "pr"
    return "unknown"


def read_shugiin() -> list[dict]:
    out: list[dict] = []
    if not SHUGIIN_XLSX.is_file():
        return out
    wb = openpyxl.load_workbook(SHUGIIN_XLSX, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        year = parse_year_from_sheet(sheet_name)
        kaiji = SHUGIIN_YEAR_TO_KAIJI.get(year) if year else None
        if year is None or kaiji is None:
            continue
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        idx = header_index(header)
        for row in it:
            name = cell(row, idx, "候補者名")
            if not name:
                continue
            ident = cell(row, idx, "識別")
            # 空 or 本番のみ（その他の識別は除外）
            if ident and ident != "本番":
                continue
            pref = normalize_pref(cell(row, idx, "ブ/県", "ブ／県"))
            district = cell(row, idx, "選挙区")
            pr_block = cell(row, idx, "比例")
            election_name = cell(row, idx, "選挙名") or f"{year}年衆院"
            kind = cell(row, idx, "選挙種類") or "衆院"
            contest = classify_contest("shugiin", district, pr_block, election_name, kind)
            out.append(
                {
                    "chamber": "shugiin",
                    "election_year": year,
                    "election_kaiji": kaiji,
                    "election_name": election_name,
                    "contest": contest,
                    "prefecture": pref,
                    "district_label": district or None,
                    "district_number": parse_district_number(district),
                    "pr_block": pr_block or None,
                    "candidate_name": name,
                    "party_short": cell(row, idx, "党派") or None,
                    "gender": cell(row, idx, "性別") or None,
                    "dob": parse_dob(row[idx["生年月日"]] if "生年月日" in idx else None),
                    "outcome": cell(row, idx, "当落") or None,
                    "ident": ident or None,
                    "source_file": SHUGIIN_XLSX.name,
                    "source_sheet": sheet_name,
                }
            )
    wb.close()
    return out


def read_sangiin(*, include_byelection: bool = False) -> list[dict]:
    out: list[dict] = []
    if not SANGIIN_XLSX.is_file():
        return out
    sheets = ["sangiin"]
    if include_byelection:
        sheets.append("補選")
    wb = openpyxl.load_workbook(SANGIIN_XLSX, read_only=True, data_only=True)
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        idx = header_index(header)
        for row in it:
            election_name = cell(row, idx, "選挙名")
            year = parse_year_from_election(election_name)
            kaiji = SANGIIN_YEAR_TO_KAIJI.get(year) if year else None
            if year is None or kaiji is None:
                continue
            name = cell(row, idx, "候補者名")
            if not name:
                continue
            ident = cell(row, idx, "識別")
            if ident and ident not in ("本番", ""):
                continue
            pref = normalize_pref(cell(row, idx, "ブ/県", "ブ／県"))
            district = cell(row, idx, "選挙区")
            pr_block = cell(row, idx, "比例")
            kind = cell(row, idx, "選挙種類") or "参院"
            contest = classify_contest("sangiin", district, pr_block, election_name, kind)
            # 参院県区は都道府県名が選挙区
            if contest == "district" and pref is None:
                continue
            out.append(
                {
                    "chamber": "sangiin",
                    "election_year": year,
                    "election_kaiji": kaiji,
                    "election_name": election_name,
                    "contest": contest,
                    "prefecture": pref,
                    "district_label": district or None,
                    "district_number": parse_district_number(district),
                    "pr_block": pr_block or None,
                    "candidate_name": name,
                    "party_short": cell(row, idx, "党派") or None,
                    "gender": cell(row, idx, "性別") or None,
                    "dob": parse_dob(row[idx["生年月日"]] if "生年月日" in idx else None),
                    "outcome": cell(row, idx, "当落") or None,
                    "ident": ident or None,
                    "source_file": SANGIIN_XLSX.name,
                    "source_sheet": sheet_name,
                }
            )
    wb.close()
    return out


def write_outputs(rows: list[dict], stamp: str) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamped = OUT_DIR / f"{stamp}_yomi_senkyodb_candidates.jsonl"
    latest = OUT_DIR / "yomi_senkyodb_candidates.jsonl"
    with stamped.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    latest.write_text(stamped.read_text(encoding="utf-8"), encoding="utf-8")

    WEB_OUT.parent.mkdir(parents=True, exist_ok=True)
    WARE_OUT.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect()
    src = stamped.as_posix().replace("'", "''")
    for dst_path in (WEB_OUT, WARE_OUT):
        dst = dst_path.as_posix().replace("'", "''")
        con.execute(
            f"""
            COPY (
              SELECT chamber,
                     election_year::BIGINT AS election_year,
                     election_kaiji::BIGINT AS election_kaiji,
                     election_name,
                     contest,
                     prefecture,
                     district_label,
                     district_number::BIGINT AS district_number,
                     pr_block,
                     candidate_name,
                     party_short,
                     gender,
                     dob,
                     outcome,
                     ident,
                     source_file,
                     source_sheet
              FROM read_json_auto('{src}', format='newline_delimited', union_by_name=true)
            ) TO '{dst}' (FORMAT PARQUET, COMPRESSION ZSTD)
            """
        )
    con.close()
    return stamped


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    rows = read_shugiin() + read_sangiin(include_byelection=False)
    stamped = write_outputs(rows, stamp)

    from collections import Counter

    by = Counter((r["chamber"], r["election_kaiji"], r["contest"]) for r in rows)
    report = OUT_DIR / f"{stamp}_normalize_report.txt"
    lines = [
        f"yomi senkyoDB normalize ({stamp})",
        f"rows={len(rows)}",
        f"web_out={WEB_OUT}",
        "",
        "counts by chamber/kaiji/contest:",
    ]
    for key, n in sorted(by.items()):
        lines.append(f"  {key}: {n}")
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(report.read_text(encoding="utf-8"))
    print(f"wrote {stamped}")
    print(f"wrote {WEB_OUT}")


if __name__ == "__main__":
    main()

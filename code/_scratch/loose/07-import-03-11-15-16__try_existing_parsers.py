# v1.1: 既存 03-11 PDF パーサの動作確認 + xls/年齢シート構造
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from soumu_election.normalize import parse_pdf_pr_elected

DATA = ROOT / "data"


def try_pdf_0311(kaiji: int) -> None:
    raw = DATA / f"shugiin{kaiji}/raw"
    paths = list(raw.glob("03-11*.pdf"))
    if not paths:
        print(f"kaiji{kaiji}: no pdf")
        return
    path = paths[0]
    doc = {
        "source_code": "03-11",
        "source_file": path.name,
        "dataset": "党派別当選人数（比例代表）",
        "election_kaiji": kaiji,
        "source_url": "",
        "pages": [],
    }
    try:
        facts = parse_pdf_pr_elected(doc, path)
        print(f"kaiji{kaiji}: facts={len(facts)}")
        metrics = {}
        for f in facts:
            metrics[f["metric"]] = metrics.get(f["metric"], 0) + 1
        print(f"  metrics={metrics}")
        sample = [f for f in facts if f["metric"] == "elected_candidates" and f.get("gender") == "total"][:6]
        for f in sample:
            print(f"  elected {f.get('pr_block')} {f.get('party')}={f.get('value')}")
    except Exception as e:
        print(f"kaiji{kaiji}: FAIL {type(e).__name__}: {e}")


def peek_xls_0311() -> None:
    import xlrd

    path = DATA / "shugiin48/raw/03-11_党派別当選人数_比例代表_000516729.xls"
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    headers = []
    for r in range(sheet.nrows):
        rowvals = [sheet.cell_value(r, c) for c in range(min(28, sheet.ncols))]
        joined = "".join(str(v) for v in rowvals)
        if "選挙区" in joined or "政党等名" in joined or "党派名" in joined:
            headers.append((r, rowvals))
    print(f"\nxls header-like rows n={len(headers)}")
    for r, vals in headers[:50]:
        print(r, vals)


def peek_0316_sheets() -> None:
    from openpyxl import load_workbook

    path = DATA / "shugiin48/raw/03-16_年齢別投票状況_000528773.xlsx"
    wb = load_workbook(path, data_only=True, read_only=True)
    print("\n03-16 sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n sheet={name}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 8:
                break
            print(" ", list(row)[:12])
    wb.close()


def peek_age_pdf_attachment() -> None:
    import pdfplumber

    base = DATA / "shugiin45/raw/03-16_45衆年齢別投票状況について_000042348_attachments"
    for path in sorted(base.glob("*.pdf")):
        print(f"\n### {path.name}")
        with pdfplumber.open(path) as pdf:
            print("pages", len(pdf.pages))
            page = pdf.pages[0]
            print((page.extract_text() or "")[:1200])
            tables = page.extract_tables() or []
            print("tables", len(tables))
            if tables:
                for ri, row in enumerate(tables[0][:12]):
                    print(f"R{ri}", row)


if __name__ == "__main__":
    for k in (44, 45, 46, 47, 49, 50, 51):
        try_pdf_0311(k)
    peek_xls_0311()
    peek_0316_sheets()
    peek_age_pdf_attachment()

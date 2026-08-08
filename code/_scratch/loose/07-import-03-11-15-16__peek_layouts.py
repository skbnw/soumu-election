# v1.0: ディレクトリ実在確認と Excel/PDF の中身覗き
from __future__ import annotations

from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"


def peek_excel(path: Path, max_rows: int = 25) -> None:
    print(f"\n### EXCEL {path}")
    if path.suffix.lower() == ".xls":
        import xlrd

        book = xlrd.open_workbook(path)
        for sheet in book.sheets():
            print(f"  sheet={sheet.name} rows={sheet.nrows} cols={sheet.ncols}")
            for r in range(min(max_rows, sheet.nrows)):
                vals = [sheet.cell_value(r, c) for c in range(min(12, sheet.ncols))]
                print(f"  R{r}: {vals}")
            break
    else:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"  sheet={name}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                print(f"  R{i}: {list(row)[:14]}")
            break
        wb.close()


def peek_pdf(path: Path, max_pages: int = 2) -> None:
    import pdfplumber

    print(f"\n### PDF {path.name}")
    with pdfplumber.open(path) as pdf:
        print(f"  pages={len(pdf.pages)}")
        for i, page in enumerate(pdf.pages[:max_pages], 1):
            text = (page.extract_text() or "")[:1200]
            print(f"  --- page {i} text ---\n{text}\n")
            tables = page.extract_tables() or []
            print(f"  tables={len(tables)}")
            if tables:
                for ri, row in enumerate(tables[0][:12]):
                    print(f"  T0R{ri}: {row}")


def main() -> None:
    for kaiji in range(44, 52):
        d = DATA / f"shugiin{kaiji}"
        print(f"\n## shugiin{kaiji}")
        for sub in ("raw", "raw_json", "normalized"):
            p = d / sub
            print(f"  {sub}: exists={p.exists()} files={len(list(p.glob('*'))) if p.exists() else 0}")

    # peek representatives
    peek_pdf(DATA / "shugiin51/raw/03-11_党派別当選人数_比例代表_001061485.pdf", 1)
    peek_pdf(DATA / "shugiin45/raw/03-11_党派別当選人数_比例代表_000037486.pdf", 1)
    peek_excel(DATA / "shugiin48/raw/03-11_党派別当選人数_比例代表_000516729.xls")
    peek_excel(DATA / "shugiin48/raw/03-15_18歳_19歳投票状況_000528771.xlsx")
    peek_excel(DATA / "shugiin48/raw/03-16_年齢別投票状況_000528773.xlsx")
    peek_pdf(DATA / "shugiin45/raw/03-16_45衆年齢別投票状況について_000042348.pdf", 2)
    # attachments for 45?
    att = DATA / "shugiin45/raw/03-16_45衆年齢別投票状況について_000042348_attachments"
    if att.exists():
        print("\n### attachments 45 03-16")
        for p in sorted(att.rglob("*")):
            if p.is_file():
                print(f"  {p.relative_to(att)} size={p.stat().st_size}")


if __name__ == "__main__":
    main()

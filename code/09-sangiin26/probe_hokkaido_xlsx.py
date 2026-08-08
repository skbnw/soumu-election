#!/usr/bin/env python3
# Download Hokkaido sangiin26 municipality xlsx and dump sheet layout
from __future__ import annotations

import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from soumu_election.download import USER_AGENT, get  # noqa: E402

OUT = ROOT / "code" / "09-sangiin26" / "output"
URLS = {
    "district": "https://www.soumu.go.jp/main_content/000828379.xlsx",
    "pr_party": "https://www.soumu.go.jp/main_content/000828381.xlsx",
    "pr_cand": "https://www.soumu.go.jp/main_content/000828382.xlsx",
}


def dump_sheet(path: Path, max_rows: int = 12, max_cols: int = 12) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"\n===== {path.name} sheets={wb.sheetnames} =====")
    for name in wb.sheetnames[:3]:
        ws = wb[name]
        print(f"--- sheet: {name} ---")
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > max_rows:
                break
            vals = list(row[:max_cols])
            print(i, vals)
    wb.close()


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers["User-Agent"] = USER_AGENT
    for key, url in URLS.items():
        dest = OUT / f"hokkaido_{key}.xlsx"
        if not dest.exists():
            resp = get(session, url)
            dest.write_bytes(resp.content)
            print("saved", dest, len(resp.content))
        else:
            print("exists", dest)
        dump_sheet(dest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
